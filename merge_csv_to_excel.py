#!/usr/bin/env python3
import argparse
import csv
import tempfile
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell


EXCEL_MAX_ROWS = 1_048_576
EXCEL_MAX_CELL_LENGTH = 32_767


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("input_path", nargs="?")
    parser.add_argument("--output")
    parser.add_argument("--self-check", action="store_true")
    return parser.parse_args()


def resolve_paths(input_path: str, output: str | None) -> tuple[Path, Path]:
    source = Path(input_path)
    if source.is_dir() and source.name == "pages":
        pages_dir = source
        run_dir = source.parent
    elif source.is_dir():
        pages_dir = source / "pages"
        run_dir = source
    else:
        raise RuntimeError(f"Input path not found: {source}")
    if not pages_dir.is_dir():
        raise RuntimeError(f"Pages directory not found: {pages_dir}")
    output_path = Path(output) if output else run_dir / f"{run_dir.name}.xlsx"
    return pages_dir, output_path


def csv_files(pages_dir: Path) -> list[Path]:
    files = sorted(pages_dir.glob("page-*.csv"))
    if not files:
        raise RuntimeError(f"No CSV page files found in {pages_dir}")
    return files


def merged_headers(files: list[Path]) -> list[str]:
    headers: list[str] = []
    seen: set[str] = set()
    for path in files:
        with path.open(newline="", encoding="utf-8-sig") as handle:
            reader = csv.reader(handle)
            try:
                current = next(reader)
            except StopIteration:
                current = []
        for column in current:
            if column not in seen:
                seen.add(column)
                headers.append(column)
    return headers


def text_cell(worksheet: object, value: str) -> WriteOnlyCell:
    cell = WriteOnlyCell(worksheet, value=value)
    cell.data_type = "s"
    cell.number_format = "@"
    return cell


def sheet_name(base: str, index: int) -> str:
    suffix = f"_{index}"
    limit = 31 - len(suffix)
    return f"{base[:limit]}{suffix}"


def write_workbook(files: list[Path], headers: list[str], output_path: Path) -> None:
    workbook = Workbook(write_only=True)
    header_index = {name: idx for idx, name in enumerate(headers)}
    sheet_count = 1
    worksheet = workbook.create_sheet(title=sheet_name(output_path.stem, sheet_count))
    worksheet.append([text_cell(worksheet, column) for column in headers])
    rows_in_sheet = 1

    for path in files:
        with path.open(newline="", encoding="utf-8-sig") as handle:
            reader = csv.DictReader(handle)
            for row in reader:
                if rows_in_sheet >= EXCEL_MAX_ROWS:
                    sheet_count += 1
                    worksheet = workbook.create_sheet(title=sheet_name(output_path.stem, sheet_count))
                    worksheet.append([text_cell(worksheet, column) for column in headers])
                    rows_in_sheet = 1

                values = [""] * len(headers)
                for key, raw in row.items():
                    if key not in header_index:
                        continue
                    value = "" if raw is None else str(raw)
                    if len(value) > EXCEL_MAX_CELL_LENGTH:
                        raise RuntimeError(f"Cell value too long for Excel in {path}: column={key}")
                    values[header_index[key]] = value
                worksheet.append([text_cell(worksheet, value) for value in values])
                rows_in_sheet += 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def run_merge(input_path: str, output: str | None) -> Path:
    pages_dir, output_path = resolve_paths(input_path, output)
    files = csv_files(pages_dir)
    headers = merged_headers(files)
    write_workbook(files, headers, output_path)
    return output_path


def write_page(path: Path, headers: list[str], rows: list[list[str]]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        writer.writerow(headers)
        writer.writerows(rows)


def self_check() -> int:
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp) / "demo"
        pages = root / "pages"
        pages.mkdir(parents=True)
        write_page(
            pages / "page-0000.csv",
            ["id", "name", "formula"],
            [["00123", "alpha", "=1+1"], ["", "beta", "00099"]],
        )
        write_page(
            pages / "page-0001.csv",
            ["name", "extra"],
            [["gamma", "A"], ["delta", "9007199254740993123"]],
        )
        output_path = run_merge(str(root), None)
        workbook = load_workbook(output_path, data_only=False)
        worksheet = workbook.worksheets[0]
        rows = list(worksheet.iter_rows(values_only=False))
        assert [cell.value for cell in rows[0]] == ["id", "name", "formula", "extra"]
        assert rows[1][0].value == "00123"
        assert rows[1][0].data_type == "s"
        assert rows[1][2].value == "=1+1"
        assert rows[1][2].data_type == "s"
        assert rows[2][0].value is None
        assert rows[3][3].value == "A"
        assert rows[4][3].value == "9007199254740993123"
    print("self-check passed")
    return 0


def main() -> int:
    args = parse_args()
    if args.self_check:
        return self_check()
    if not args.input_path:
        raise RuntimeError("input_path is required unless --self-check is used")
    output_path = run_merge(args.input_path, args.output)
    print(f"wrote {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
